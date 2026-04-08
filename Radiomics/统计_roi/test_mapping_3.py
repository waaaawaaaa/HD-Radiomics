# 对huntington数据的统计，三类，HC PRE_HD HD做统计

import os
import pandas as pd
import numpy as np
from scipy.stats import shapiro, levene, f_oneway, kruskal
import scikit_posthocs as sp
from statsmodels.stats.multicomp import pairwise_tukeyhsd
from openpyxl import Workbook
from openpyxl import load_workbook


# 定义加粗函数
def bold_p_values(val):
    """
    如果 p 值小于 0.05，则加粗显示。
    """
    try:
        if float(val) < 0.05:
            return 'font-weight: bold'
        else:
            return ''
    except ValueError:
        return ''

# 读取数据
seq = 'T2Mapping'  # T2Mapping   T2star_Mapping
excel = '腐蚀1_T2_cx'
file_root = 'H:/重建数据/mapping_unet_best/peizhun/PD_out_t2m/ROI_processed/'
filename = os.path.join(file_root, f'{seq}_{excel}.xlsx')
data = pd.read_excel(filename)

# 输出文件
output_filename = f'test3_{seq}_{excel}2.xlsx'
excel_output = os.path.join(file_root, output_filename)

# 获取唯一 ROI
roi_values = data['ROI'].unique()

# 如果文件已存在，则删除
if os.path.exists(excel_output):
    os.remove(excel_output)

wb = Workbook()
wb.save(excel_output)

all_comparisons = []

# 处理不同 ROI
for roi in roi_values:
    roi_data = data[data['ROI'] == roi].copy()
    roi_data.drop(columns=['ROI'], inplace=True)

    # 按 'ClassicValue' 排序
    sorted_table = roi_data.sort_values(by='ClassicValue')
    features_name = sorted_table.columns[:-4]

    mean_sd = []
    # 创建一个空列表，用于存储所有 ROI 的 comparisons_data


    for feature in features_name:
        feature_HC = sorted_table.loc[sorted_table['ClassicValue'] == 0, feature]
        feature_pre_HD = sorted_table.loc[sorted_table['ClassicValue'] == 1, feature]
        feature_HD = sorted_table.loc[sorted_table['ClassicValue'] == 2, feature]

        # 均值和标准差
        mean_sd.append([
            feature_HC.mean(), feature_HC.std(),
            feature_pre_HD.mean(), feature_pre_HD.std(),
            feature_HD.mean(), feature_HD.std()
        ])

        # 数据合并
        data_all = np.concatenate([feature_HC, feature_pre_HD, feature_HD])
        group = np.concatenate([
            np.full(len(feature_HC), 1),  # HC 组
            np.full(len(feature_pre_HD), 2),  # pre-HD 组
            np.full(len(feature_HD), 3)  # HD 组
        ])

        # 正态性检验（Shapiro-Wilk）
        p_shapiro_HC = shapiro(feature_HC).pvalue if len(feature_HC) >= 3 else 1
        p_shapiro_pre_HD = shapiro(feature_pre_HD).pvalue if len(feature_pre_HD) >= 3 else 1
        p_shapiro_HD = shapiro(feature_HD).pvalue if len(feature_HD) >= 3 else 1

        # 方差齐性检验（Levene）
        p_levene = levene(feature_HC, feature_pre_HD, feature_HD).pvalue  #正态数据看基于平均值，偏态数据看基于中位数，默认中位数

        # 选择统计检验方法
        if p_shapiro_HC > 0.05 and p_shapiro_pre_HD > 0.05 and p_shapiro_HD > 0.05 and p_levene > 0.05:

            p_test = f_oneway(feature_HC, feature_pre_HD, feature_HD).pvalue
            test_type = 1
            # # 执行 Tukey HSD 检验
            if p_test < 0.05:
                tukey_result = pairwise_tukeyhsd(data_all, group)
                comparisons = [tukey_result.pvalues[0], tukey_result.pvalues[1], tukey_result.pvalues[2]]
            else:
                comparisons = ['-', '-', '-']

        else:
            # Kruskal-Wallis（非参数检验）
            p_test = kruskal(feature_HC, feature_pre_HD, feature_HD).pvalue
            test_type = 2

            # 事后检验（Dunn）
            if p_test < 0.05:
                dunn_results = sp.posthoc_dunn([feature_HC, feature_pre_HD, feature_HD], p_adjust='bonferroni')
                comparisons = [dunn_results.loc[1, 2], dunn_results.loc[1, 3], dunn_results.loc[2, 3]]
            else:
                comparisons = ['-', '-', '-']

        mean_sd[-1].extend([p_shapiro_HC, p_shapiro_pre_HD, p_shapiro_HD, p_levene, test_type, p_test] + comparisons)

    # 转换为 DataFrame 并保存
    result_table = pd.DataFrame(mean_sd, columns=['mean_hc', 'sd_hc', 'mean_pre_hd', 'sd_pre_hd', 'mean_hd', 'sd_hd',
                                                  'p_shapiro_HC', 'p_shapiro_pre_HD', 'p_shapiro_HD', 'p_levene',
                                                  'test_type_1anova_2kruskal', 'p_test', 'hc_vs_pre_hd', 'hc_vs_hd', 'pre_hd_vs_hd'])

    # 应用样式
    styled_result_table = result_table.style.map(
        bold_p_values,
        subset=['hc_vs_pre_hd', 'hc_vs_hd', 'pre_hd_vs_hd']  # 只对 comparisons 列应用样式
    )

    # 提取数据部分（去掉 'subject' 列）
    data_sorted = sorted_table.iloc[:, :-1]  # 除去最后一列（subject）

    # 转置数据并创建新 DataFrame
    data_table = data_sorted.T
    data_table.columns = sorted_table['Subject']  # 使用 'Subject' 列作为列名
    data_table['ROI'] = data_table.index  # 将原行名（变量名）作为 'ROI' 列

    # 重新排列列顺序，使 'ROI' 变为第一列
    cols = ['ROI'] + [col for col in data_table.columns if col != 'ROI']
    data_table = data_table[cols]

    # 提取最后三列
    comparisons_data = result_table.iloc[:, -3:]
    all_comparisons.append(comparisons_data)

    # 打开 Excel 进行写入，不要每次都新建文件
    with pd.ExcelWriter(excel_output, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        data_table.to_excel(writer, sheet_name=str(roi), startrow=0, startcol=0, index=False)
        styled_result_table.to_excel(writer, sheet_name=str(roi), startrow=0, startcol=data_table.shape[1] + 1, index=False)  # 从第 64 列开始写入

    print(f'统计结果和数据表已成功保存到 {excel_output} 的 {str(roi)} 表单中！')

# 假设 all_comparisons 是一个包含多个 DataFrame 的列表
all_comparisons_df = pd.concat(all_comparisons, axis=1)
# 提取 data_table 的 ROI 列
roi_column = data_table['ROI'].reset_index(drop=True)
# 将提取的 ROI 列添加到 all_comparisons_df 的最左边
all_comparisons_df.insert(0, 'ROI', roi_column)
# 将数据四舍五入到 4 位小数
all_comparisons_df = all_comparisons_df.round(4)

# 保存到 Excel
with pd.ExcelWriter(excel_output, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
    all_comparisons_df.to_excel(writer, sheet_name='All_ROI_Comparisons',startrow=1, startcol=0, index=False)
print(f'统计结果和数据表已成功保存到 {excel_output} 的 All_ROI_Comparisons 表单中！')

# 加载 Excel 文件
wb = load_workbook(excel_output)

# 检查并删除空的 sheet
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(1, 1).value is None:
        wb.remove(sheet)

# 保存修改后的 Excel 文件
wb.save(excel_output)


